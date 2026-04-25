"""
Crea el Excel maestro de entrada con 50 estudiantes de ejemplo
y el script de procesamiento masivo.
"""
import random
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
import os

OUT = "/sessions/awesome-modest-lovelace/mnt/RETO 2- BECA EAFIT- SER ANDI/demo"

# ─── Datos ficticios pero realistas ─────────────────────────────────────────
NOMBRES = ["VALENTINA","SANTIAGO","MARÍA CAMILA","JUAN DAVID","DANIELA",
    "SEBASTIÁN","LAURA","ANDRÉS FELIPE","ISABELLA","MIGUEL ANGEL",
    "SOFIA","CARLOS ALBERTO","NATALIA","ALEJANDRO","PAULA ANDREA",
    "NICOLÁS","SARA","TOMÁS","JULIANA","ESTEBAN","GABRIELA","DAVID",
    "MANUELA","MATEO","ANA MARÍA","SAMUEL","LUISA FERNANDA","JORGE",
    "CATALINA","RICARDO","MARIANA","LUIS MIGUEL","DIANA","PABLO",
    "CAROLINA","FERNANDO","JESSICA","IVÁN","VERÓNICA","RAFAEL",
    "MÓNICA","CHRISTIAN","TATIANA","SERGIO","MELISSA","VICTOR",
    "PATRICIA","GABRIEL","VANESSA","ALEJANDRA"]

APELLIDOS = ["GÓMEZ RÍOS","MARTÍNEZ López","GARCÍA HERRERA","RODRÍGUEZ OSORIO",
    "FERNÁNDEZ CASTRO","DÍAZ MORALES","TORRES JIMÉNEZ","VARGAS SÁNCHEZ",
    "MORENO RESTREPO","RAMÍREZ FRANCO","PÉREZ AGUILAR","ÁLVAREZ CANO",
    "RUIZ ZAPATA","SÁNCHEZ MEJÍA","HENAO MONTOYA","OSPINA GIRALDO",
    "CARDONA VÉLEZ","MESA ARANGO","SALAZAR BEDOYA","ARANGO TOBÓN",
    "VALENCIA CORREA","RÍOS ZULUAGA","CANO LONDOÑO","GÓMEZ POSADA",
    "HERRERA MUÑOZ","ESCOBAR VILLA","BUSTAMANTE SIERRA","ACOSTA PARRA",
    "CASTAÑO MONTES","AGUIRRE DUQUE"]

PROGRAMAS = [
    "Administración de Empresas","Ingeniería de Sistemas","Ingeniería Industrial",
    "Negocios Internacionales","Comunicación Social","Derecho","Contaduría Pública",
    "Psicología","Economía","Ingeniería Civil","Marketing","Finanzas",
    "Diseño Gráfico","Arquitectura","Ingeniería Biomédica","Música",
    "Ciencias Políticas","Filosofía","Geología","Matemáticas Aplicadas"
]

EMPRESAS = [
    ("BANCOLOMBIA S.A.","S.A.","890.903.938-8","Medellín, Antioquia","CARLOS MARIO AGUDELO VILLA","72.001.234","Vicepresidente de Talento"),
    ("GRUPO NUTRESA S.A.S.","S.A.S.","890.100.324-1","Medellín, Antioquia","PATRICIA DURÁN CANO","43.876.543","Directora de RRHH"),
    ("EPM S.A. E.S.P.","S.A.","890.905.762-2","Medellín, Antioquia","JORGE IVÁN VÉLEZ","71.345.678","Gerente de Personas"),
    ("SURA ASSET MANAGEMENT","S.A.","811.002.349-3","Medellín, Antioquia","DIANA MARCELA RÍOS","43.654.321","Jefe de Talento"),
    ("TECNOLOGÍA INNOVADORA S.A.S.","S.A.S.","900.123.456-7","Medellín, Antioquia","CARLOS PÉREZ GÓMEZ","71.234.567","Gerente General"),
    ("SOFASA S.A.","S.A.","890.907.795-5","Envigado, Antioquia","ANDRÉS FELIPE RESTREPO","71.456.789","Director Operaciones"),
    ("GRUPO ÉXITO S.A.","S.A.","860.002.528-9","Medellín, Antioquia","LUISA BERMÚDEZ ARANGO","42.987.654","Gerente Talento"),
    ("COMFAMA","Fundación","890.903.922-6","Medellín, Antioquia","MAURICIO RESTREPO VILLA","71.567.890","Coordinador RRHH"),
    ("KONECTA BTO S.A.S.","S.A.S.","900.456.789-2","Medellín, Antioquia","VERÓNICA SALAZAR MESA","43.234.567","Directora de Operaciones"),
    ("INTERNEXA S.A. E.S.P.","S.A.","830.030.619-8","Medellín, Antioquia","HERNÁN CORREA GIRALDO","71.678.901","Gerente de Personas"),
]

TUTORES = [
    ("JUAN DAVID SALAZAR MESA","98.765.432","Coordinador de Proyectos"),
    ("MARCELA RESTREPO GIRALDO","43.234.567","Jefe de Área"),
    ("ANDRÉS VARGAS OSPINA","71.890.123","Director de Tecnología"),
    ("CAROLINA HENAO CANO","42.345.678","Supervisora de Prácticas"),
    ("SERGIO MONTOYA ARANGO","98.456.789","Líder de Equipo"),
    ("DIANA RÍOS ZULUAGA","43.567.890","Coordinadora Académica"),
    ("FELIPE GÓMEZ VÉLEZ","71.123.456","Gerente de Área"),
    ("NATALIA CARDONA MESA","43.789.012","Jefe de Talento"),
]

MONITORES_EAFIT = [
    ("LAURA VANESSA CATAÑO PULGARÍN","43.111.222"),
    ("PABLO ANDRÉS MESA GIRALDO","71.333.444"),
    ("SANDRA MILENA RÍOS OSPINA","43.555.666"),
    ("CARLOS EDUARDO VÉLEZ ARANGO","98.777.888"),
]

ACTIVIDADES_BASE = {
    "Administración de Empresas": [
        "Apoyar la gestión administrativa y operativa del área asignada.",
        "Elaborar informes de gestión y reportes de indicadores de desempeño.",
        "Participar en reuniones de seguimiento y documentar acuerdos.",
        "Apoyar procesos de mejora continua y gestión de proyectos.",
        "Colaborar en la atención y seguimiento a clientes internos y externos.",
        "Sistematizar información y mantener bases de datos actualizadas.",
    ],
    "Ingeniería de Sistemas": [
        "Apoyar el desarrollo y mantenimiento de aplicaciones de software.",
        "Participar en pruebas de calidad y documentación técnica.",
        "Colaborar en el análisis de requerimientos del sistema.",
        "Apoyar la gestión de infraestructura tecnológica.",
        "Contribuir al desarrollo de módulos del sistema de información.",
        "Realizar documentación de procesos y manuales técnicos.",
    ],
    "default": [
        "Apoyar el desarrollo de las actividades del área de práctica.",
        "Participar activamente en proyectos y reuniones del equipo.",
        "Elaborar informes y reportes solicitados por el tutor.",
        "Documentar procesos, procedimientos y lecciones aprendidas.",
        "Colaborar en la implementación de mejoras en los procesos.",
        "Mantener comunicación permanente con tutor y monitor.",
    ]
}

random.seed(42)

def fecha_aleatoria_inicio():
    base = date(2025, 2, 1)
    dias = random.randint(0, 180)
    d = base + timedelta(days=dias)
    # Primer día hábil del mes siguiente
    return d.replace(day=random.choice([1, 5, 10, 15]))

def monto_aleatorio():
    montos = [
        ("UN MILLÓN", "1.000.000"),
        ("UN MILLÓN CIENTO CINCUENTA MIL", "1.150.000"),
        ("UN MILLÓN TRESCIENTOS MIL", "1.300.000"),
        ("UN MILLÓN QUINIENTOS MIL", "1.500.000"),
        ("DOS MILLONES", "2.000.000"),
    ]
    return random.choice(montos)

def get_actividades(programa):
    acts = ACTIVIDADES_BASE.get(programa, ACTIVIDADES_BASE["default"])
    return acts[:4]

# ─── Crear 50 registros ──────────────────────────────────────────────────────
registros = []
for i in range(50):
    nombre = f"{NOMBRES[i]} {APELLIDOS[i % len(APELLIDOS)]}"
    cedula = f"{random.randint(1000,1100)}.{random.randint(100,999)}.{random.randint(100,999)}"
    programa = random.choice(PROGRAMAS)
    empresa = EMPRESAS[i % len(EMPRESAS)]
    tutor = TUTORES[i % len(TUTORES)]
    monitor = MONITORES_EAFIT[i % len(MONITORES_EAFIT)]
    remunerada = random.random() > 0.25  # 75% remuneradas
    quien_arl = "Organización" if random.random() > 0.3 else "EAFIT"
    monto = monto_aleatorio() if remunerada else ("", "")
    fi = fecha_aleatoria_inicio()
    ff = fi + timedelta(days=random.choice([150, 180, 210, 240]))
    acts = get_actividades(programa)

    registros.append({
        "tipo_experiencia": "Práctica",
        "remunerada": "Sí" if remunerada else "No",
        "quien_paga_arl": quien_arl,
        "nombre_empresa": empresa[0],
        "tipo_sociedad": empresa[1],
        "nit_empresa": empresa[2],
        "ciudad_empresa": empresa[3],
        "nombre_representante": empresa[4],
        "cedula_representante": empresa[5],
        "cargo_representante": empresa[6],
        "nombre_estudiante": nombre,
        "cedula_estudiante": cedula,
        "programa_academico": programa,
        "fecha_inicio": fi.strftime("%d/%m/%Y"),
        "fecha_fin": ff.strftime("%d/%m/%Y"),
        "monto_letras": monto[0],
        "monto_numeros": monto[1],
        "nombre_tutor": tutor[0],
        "cedula_tutor": tutor[1],
        "cargo_tutor": tutor[2],
        "nombre_monitor": monitor[0],
        "cedula_monitor": monitor[1],
        "actividad_1": acts[0],
        "actividad_2": acts[1],
        "actividad_3": acts[2],
        "actividad_4": acts[3],
        "fecha_firma": fi.strftime("%d/%m/%Y"),
    })

# ─── Construir Excel con openpyxl ────────────────────────────────────────────
wb = Workbook()

# ── Hoja 1: DATOS ────────────────────────────────────────────────────────────
ws = wb.active
ws.title = "Datos"

AZUL_OSCURO = "003087"
AZUL_CLARO  = "D6E4F7"
VERDE       = "E2F0D9"
GRIS        = "F2F2F2"
BLANCO      = "FFFFFF"

COLUMNAS = [
    ("tipo_experiencia",   "Tipo de Experiencia",        18),
    ("remunerada",         "¿Remunerada?",               14),
    ("quien_paga_arl",     "¿Quién paga ARL?",           16),
    ("nombre_empresa",     "Nombre Empresa",              30),
    ("tipo_sociedad",      "Tipo Sociedad",               12),
    ("nit_empresa",        "NIT",                         16),
    ("ciudad_empresa",     "Ciudad Empresa",              22),
    ("nombre_representante","Representante Legal",        30),
    ("cedula_representante","Cédula Rep. Legal",          16),
    ("cargo_representante","Cargo Rep. Legal",            22),
    ("nombre_estudiante",  "Nombre Estudiante",           30),
    ("cedula_estudiante",  "Cédula Estudiante",           18),
    ("programa_academico", "Programa Académico",          28),
    ("fecha_inicio",       "Fecha Inicio (DD/MM/AAAA)",   22),
    ("fecha_fin",          "Fecha Fin (DD/MM/AAAA)",      22),
    ("monto_letras",       "Monto en Letras",             30),
    ("monto_numeros",      "Monto en Números",            18),
    ("nombre_tutor",       "Nombre Tutor (Empresa)",      30),
    ("cedula_tutor",       "Cédula Tutor",                16),
    ("cargo_tutor",        "Cargo Tutor",                 22),
    ("nombre_monitor",     "Nombre Monitor (EAFIT)",      30),
    ("cedula_monitor",     "Cédula Monitor",              16),
    ("actividad_1",        "Actividad 1",                 45),
    ("actividad_2",        "Actividad 2",                 45),
    ("actividad_3",        "Actividad 3",                 45),
    ("actividad_4",        "Actividad 4",                 45),
    ("fecha_firma",        "Fecha Firma (DD/MM/AAAA)",    22),
]

# Encabezado fila 1: título
ws.merge_cells("A1:AA1")
c = ws["A1"]
c.value = "TALENTO EAFIT — CARGA MASIVA DE CONVENIOS"
c.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
c.fill = PatternFill("solid", fgColor=AZUL_OSCURO)
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 28

# Encabezado fila 2: columnas
thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col_idx, (key, label, width) in enumerate(COLUMNAS, start=1):
    cell = ws.cell(row=2, column=col_idx, value=label)
    cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="0057A8")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
    ws.column_dimensions[get_column_letter(col_idx)].width = width
ws.row_dimensions[2].height = 32

# Datos
for row_idx, reg in enumerate(registros, start=3):
    bg = BLANCO if row_idx % 2 == 1 else GRIS
    for col_idx, (key, label, width) in enumerate(COLUMNAS, start=1):
        val = reg.get(key, "")
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = Font(name="Arial", size=9)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(vertical="center", wrap_text=False)
        cell.border = border

# Freeze header
ws.freeze_panes = "A3"

# ── Hoja 2: INSTRUCCIONES ────────────────────────────────────────────────────
wi = wb.create_sheet("Instrucciones")
instrucciones = [
    ("INSTRUCCIONES DE USO", True, AZUL_OSCURO, "FFFFFF", 22),
    ("", False, BLANCO, "000000", 11),
    ("1. COMPLETAR LOS DATOS", True, "0057A8", "FFFFFF", 13),
    ("   Llene cada fila con los datos de un estudiante. No elimine columnas.", False, BLANCO, "000000", 11),
    ("   Cada fila = 1 convenio a generar.", False, BLANCO, "000000", 11),
    ("", False, BLANCO, "000000", 11),
    ("2. CAMPOS OBLIGATORIOS", True, "0057A8", "FFFFFF", 13),
    ("   Todos los campos son obligatorios.", False, BLANCO, "000000", 11),
    ("   Si la práctica NO es remunerada, dejar 'Monto en Letras' y 'Monto en Números' vacíos.", False, BLANCO, "000000", 11),
    ("", False, BLANCO, "000000", 11),
    ("3. FORMATO DE FECHAS", True, "0057A8", "FFFFFF", 13),
    ("   Usar formato DD/MM/AAAA   Ejemplo: 05/05/2025", False, BLANCO, "000000", 11),
    ("", False, BLANCO, "000000", 11),
    ("4. CAMPO 'REMUNERADA'", True, "0057A8", "FFFFFF", 13),
    ("   Escribir exactamente: Sí   o   No", False, BLANCO, "000000", 11),
    ("", False, BLANCO, "000000", 11),
    ("5. CAMPO 'QUIÉN PAGA ARL'", True, "0057A8", "FFFFFF", 13),
    ("   Escribir exactamente: Organización   o   EAFIT", False, BLANCO, "000000", 11),
    ("", False, BLANCO, "000000", 11),
    ("6. GENERAR LOS CONVENIOS", True, "0057A8", "FFFFFF", 13),
    ("   En la app Streamlit, ir a la pestaña 'Carga Masiva'.", False, BLANCO, "000000", 11),
    ("   Subir este archivo Excel y hacer clic en 'Generar todos los convenios'.", False, BLANCO, "000000", 11),
    ("   Los documentos se descargarán en un archivo .zip con todos los convenios.", False, BLANCO, "000000", 11),
    ("", False, BLANCO, "000000", 11),
    ("CONTACTO SOPORTE", True, "16a34a", "FFFFFF", 13),
    ("   Equipo Reto 2 · Beca IA EAFIT · SER ANDI 2025", False, BLANCO, "000000", 11),
]

wi.column_dimensions["A"].width = 80
for r_idx, (texto, bold, bg, fg, size) in enumerate(instrucciones, start=1):
    cell = wi.cell(row=r_idx, column=1, value=texto)
    cell.font = Font(name="Arial", bold=bold, size=size, color=fg)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(vertical="center", indent=1)
    wi.row_dimensions[r_idx].height = 22 if texto else 8

out_path = os.path.join(OUT, "TALENTO_EAFIT_Carga_Masiva.xlsx")
wb.save(out_path)
print(f"✅ Excel creado: {out_path}")
print(f"   {len(registros)} estudiantes registrados")
